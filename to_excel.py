import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
now = datetime.now()


def create_excel(ppe):

    now = datetime.now()
    date = now.strftime("%x")
    time = now.strftime("%H:%M:%S")
    manager = "Joshua Smith"
    location = 'Factory 1'
    camera = 1
    filename = "logs.xlsx"
    ppe_class = round(ppe.item())
    

    df = pd.DataFrame({
            'Date': [date],'Time': [time],'Manager': [manager], 
                        'Violation Type': [ppe_class], 'Location': [location], 'Camera': [camera]})
    

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a',if_sheet_exists='overlay')
    writer.workbook = openpyxl.load_workbook(filename)
    df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False, header= False)
    writer.close()
 


    
   




